from django.core.management.base import BaseCommand

from openpyxl import load_workbook
from aiogram import Bot, Dispatcher, executor, types
from datetime import datetime
import requests
from fake_useragent import UserAgent
from user.models import User

from asgiref.sync import sync_to_async

# from main.models import Setting



class Command(BaseCommand):
	def handle(self, *args, **options):
		# settings = Setting.get_settings()
		# token = settings.token
		token = '5719307109:AAFaZPM8NntlByxUW4Co81ia1Hs9ikG3CQE'
		bot = Bot(token=token)
		dp = Dispatcher(bot)

		merchant_id = []


		@sync_to_async
		def get_user(message):
			user = User.objects.filter(telegram_id=message.chat.id)
			if user:
				user = user[0]
			return user


		@dp.message_handler(commands=['start'])
		async def start(message):
			user = await get_user(message)

			if user and user.check_active():
			# if 1 == 2:
				update_price = types.KeyboardButton('⬇️ Загрузить прайслист')
				markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
				markup.row(update_price)

				text = "<b>Чтобы воспользоваться ботом загрузите ваш прайслист!</b>"
				await bot.send_message(message.chat.id, text, parse_mode='HTML', reply_markup=markup)
			else:
				text = "<b>У вас нет доступа к использованию бота.</b>"
				await bot.send_message(message.chat.id, text, parse_mode='HTML')


		@dp.message_handler(content_types=['text'])
		async def get_text(message):
			user = await get_user(message)

			if user and user.check_active():
			# if 1 == 2:
				if message.text == '⬇️ Загрузить прайслист':
					await bot.send_message(message.chat.id, 'Отправьте прайслист в формате Exel таблицы')


		@dp.message_handler(content_types=['document'])
		async def download_pricelist(message):
			user = await get_user(message)

			if user and user.check_active():
			# if 1 == 2:
			# try:
				date = datetime.now().strftime("%d_%m_%Y %H_%M_%S")
				file_info = await bot.get_file(message.document.file_id)
				downloaded_file = await bot.download_file(file_info.file_path)
				with open(f'main/management/commands/pricelists/{date}.xlsx', 'wb') as new_file:
					new_file.write(downloaded_file.getvalue())

				products = await get_products_with_file(f'main/management/commands/pricelists/{date}.xlsx')
				message_wait = await bot.send_message(message.chat.id, '⏰ Пожалуйста ожидайте...')

				i = 1
				for product in products:
					await bot.edit_message_text(chat_id=message.chat.id, message_id=message_wait.message_id, text=f'⏰ Пожалуйста ожидайте... Прогресс {i}/{len(products)}')
					product = await parse_product(product)
					i += 1

				await bot.send_message(message.chat.id, 'Идет сохранение данных...')
				await save_products_with_file(f'main/management/commands/pricelists/{date}.xlsx', products)

				f = open(f'main/management/commands/pricelists/{date}.xlsx', "rb")
				await bot.send_document(message.chat.id, f)
				f.close()
			# except:
			# 	await bot.send_message(message.chat.id, 'Ошибка при обработке данных')


		async def save_products_with_file(file, products):
			wb = load_workbook(filename=file)
			ws = wb.active

			for product in products:
				ws.cell(row=product['row'], column=4, value=product['price'])
			
			wb.save(file)


		async def get_products_with_file(file):
			wb = load_workbook(filename=file)
			ws = wb.active

			products = []
			for row in range(2, ws.max_row + 1):
				price = 0
				if ws.cell(row, 4).value:
					price = int(ws.cell(row, 4).value)
				product = {
					"row": row,
					"price": price,
					"minimum_price": ws.cell(row, 6).value,
					"article": ws.cell(row, 7).value,
				}
				products.append(product)
			
			return products


		async def parse_product(product):
			ua = UserAgent()
			user_agent = str(ua.random)
			headers = {
				'Accept': 'application/json, text/*',
				'Accept-Encoding': 'gzip, deflate, br',
				'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
				'Connection': 'keep-alive',
				'Content-Length': '267',
				'Content-Type': 'application/json; charset=UTF-8',
				'Host': 'kaspi.kz',
				'Origin': 'https://kaspi.kz',
				'Referer': 'https://kaspi.kz/',
				'User-Agent': user_agent,
			}
			data = {
				'cityId': "750000000",
				'id': f"{product['article']}",
				'limit': 5,
				'merchantUID': None,
				'page': 0,
				'sort': True,
			}

			response = requests.post(f'https://kaspi.kz/yml/offer-view/offers/{product["article"]}', json=data, headers=headers).json()

			try:
				minimum_seller = response['offers'][0]
				minimum_price = int(minimum_seller['price'])

				if str(minimum_seller['merchantId']) in merchant_id:
					try:
						minimum_seller = response['offers'][1]
						minimum_price = int(minimum_seller['price'])
					except IndexError:
						product['price'] = minimum_price
						print('------------- Вы единственный продавец! -------------')
						return product

			except (IndexError, KeyError):
				error = f'------------- Минимальная цена товара {product["article"]} не найдена! -------------'
				print(error)
				return product

			minimum_price = minimum_price - 2
			if minimum_price < product['minimum_price'] and product['minimum_price'] != 0:
				minimum_price = product['minimum_price']

			product['price'] = minimum_price

			print(f'------------- Товар с артикулом {product["article"]} пропаршен! -------------')
			return product


		executor.start_polling(dp, skip_updates=True)
