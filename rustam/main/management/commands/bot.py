from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils import timezone

import os
import openpyxl
from fake_useragent import UserAgent
import requests
import time
import telebot

from user.models import User

# from selenium import webdriver
from seleniumwire import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options


PROXY_STRING = 'aishanishopkz:Y8wTrLpxsC@212.116.246.66:50100'


def save_products_with_file(file, products):
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.active

    for product in products:
        if product['price']:
            ws.cell(row=product['row'], column=4, value=product['price'])

    wb.save(file)


def get_products_with_file(file):
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.active

    products = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 5).value == 'no':
            continue
        price = 0
        if ws.cell(row, 4).value:
            price = int(ws.cell(row, 4).value)

        minimum_price = ws.cell(row, 6).value
        if minimum_price:
            minimum_price = int(minimum_price)
        else:
            minimum_price = 0

        product = {
            "row": row,
            "price": price,
            "minimum_price": minimum_price,
            "article": ws.cell(row, 7).value,
        }
        if price:
            products.append(product)

    return products


def parse_product(product, merchant_id):
    ua = UserAgent()
    user_agent = str(ua.random)

    url = f'https://kaspi.kz/yml/offer-view/offers/{product["article"]}'
    headers = {
        'Referer': 'https://kaspi.kz/',
        'User-Agent': user_agent,
    }
    data = {
        'cityId': "750000000",
        'id': f"{product['article']}",
        'limit': 5,
        'merchantUID': '',
        'page': 0,
        'sort': True,
    }
    proxies = {
        'https': f'http://{PROXY_STRING}'
    }

    response = requests.post(
        url, 
        headers=headers, 
        json=data,
        proxies=proxies
    ).json()

    try:
        minimum_seller = response['offers'][0]
        minimum_price = int(minimum_seller['price'])

        if str(minimum_seller['merchantId']) in merchant_id:
            try:
                minimum_seller = response['offers'][1]
                minimum_price = int(minimum_seller['price'])
            except IndexError:
                product['price'] = minimum_price
                print('------------- Вы единственный продавец! -------------')
                return product

    except (IndexError, KeyError):
        error = f'------------- Минимальная цена товара {product["article"]} не найдена! -------------'
        print(error)
        return product

    minimum_price = minimum_price - 2
    if minimum_price < product['minimum_price'] and product['minimum_price'] != 0:
        minimum_price = product['minimum_price']

    product['price'] = minimum_price

    print(f'------------- Товар с артикулом {product["article"]} пропаршен! -------------')
    return product


def login(login_email, login_password):
    print('Логиним')
    chrome_options = Options()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--headless")

    proxy_options = {
        "proxy": {
            "https": f"http://{PROXY_STRING}"
        }
    }

    driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chrome_options, seleniumwire_options=proxy_options)

    time.sleep(3)
    driver.get('https://kaspi.kz/merchantcabinet/login?logout=true')
    print('Зашли на страницу логина')

    time.sleep(1)
    driver.find_element(by=By.CSS_SELECTOR, value='input#email').send_keys(login_email)
    time.sleep(1)
    driver.find_element(by=By.CSS_SELECTOR, value='input#password').send_keys(login_password)
    time.sleep(1)
    driver.find_element(by=By.CSS_SELECTOR, value='button[type="submit"]').click()
    print('Ввели данные')

    time.sleep(5)
    try:
        if driver.find_element(by=By.CSS_SELECTOR, value='#main-nav-offers'):
            pass
    except NoSuchElementException:
        return login(login_email, login_password)

    return driver


def submit_pricelist(user, file):
    driver = login(user.kaspi_login, user.kaspi_password)
    print('Залогинились')

    driver.find_element(by=By.CSS_SELECTOR, value='#main-nav-offers').click()
    time.sleep(2)
    driver.find_elements(by=By.CSS_SELECTOR, value='.page-heading__controls button')[1].click()
    try:
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[for="autoUploadId"]')))
    except:
        return None
    time.sleep(2)
    print('Загружаем ссылку')

    driver.find_element(by=By.CSS_SELECTOR, value='[for="autoUploadId"]').click()
    time.sleep(2)

    driver.find_element(by=By.CSS_SELECTOR, value='[name="filePathHttp"]').clear()
    time.sleep(1)
    driver.find_element(by=By.CSS_SELECTOR, value='[name="filePathHttp"]').send_keys('https://www.price-bot.kz' + user.pricelist_xml.url)
    # driver.find_element(by=By.CSS_SELECTOR, value='[name="filePathHttp"]').send_keys('https://price-bot.kz/media/pricelists_xml/2022_09_12_12_53_30.xml')
    time.sleep(2)

    driver.find_element(by=By.CSS_SELECTOR, value='.button').click()
    time.sleep(2)
    try:
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, '._secondary.button')))
    except:
        return None
    try:
        driver.find_elements(by=By.CSS_SELECTOR, value='.button')[1].click()
    except:
        return None
    print('Сохранили')
    time.sleep(10)
    driver.close()


class Command(BaseCommand):
    def handle(self, *args, **options):
        only_save = False
        while True:
            users = User.objects.all().order_by('id')
            # users = User.objects.filter(email="nurbol_15@mail.ru").order_by('id')

            for user in users:
                print(f'Проходим {user.username}')
                if user.active_date < timezone.now() or not user.pricelist:
                    continue
                file = os.path.join(settings.MEDIA_ROOT, user.pricelist.name)

                try:
                    if not only_save:
                        products = get_products_with_file(file)

                        i = 1
                        for product in products:
                            print(f'Парсим {i} товар')
                            product = parse_product(product, user.get_skip_merchant_id())
                            i += 1
                        print('Сохранили товар')
                        save_products_with_file(file, products)
                        user.save()

                    bot = telebot.TeleBot("5719307109:AAFaZPM8NntlByxUW4Co81ia1Hs9ikG3CQE")
                    try:
                        text = f'Ваш прайслист обновлён\nСкачать: https://www.price-bot.kz{user.pricelist_xlsx.url}'
                        bot.send_message(user.telegram_id, text, parse_mode='HTML')
                    except:
                        print('Человек не привязал свой id телеграм')
                except:
                    try:
                        text = f'У вас загружен прайс в неверном формате!'
                        bot.send_message(user.telegram_id, text, parse_mode='HTML')
                    except:
                        print('Человек не привязал свой id телеграм')

                try:
                    if user.kaspi_login and user.kaspi_password:
                        if user.pricelist_xml:
                            submit_pricelist(user, file)
                    print(f'Закончили {user.username}')
                except:
                    pass

            print('Отдыхаем 1 час')
            time.sleep(3600)
