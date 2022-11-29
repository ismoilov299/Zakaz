import pandas
from django.contrib.auth.models import AbstractUser
from django.db import models
from django.utils import timezone


import openpyxl
import xml.etree.ElementTree as ET
import os
from django.conf import settings
from django.core.files import File
import shutil


class User(AbstractUser):
    telegram_id = models.CharField('Telegram Id', max_length=100, default='', blank=True)
    pricelist = models.FileField('Прайслит', upload_to='pricelists', default='', blank=True)
    pricelist_xlsx = models.FileField('Прайслит XLSX', upload_to='pricelists_xlsx', default='', blank=True)
    pricelist_xml = models.FileField('Прайслит XML', upload_to='pricelists_xml', default='', blank=True)
    active_date = models.DateTimeField('Дата активности', default=timezone.now)
    active = models.BooleanField('Активный', default=False)

    kaspi_login = models.CharField('Логин Kaspi', default='', max_length=100, blank=True)
    kaspi_password = models.CharField('Пароль Kaspi', default='', max_length=100, blank=True)
    kaspi_company = models.CharField('Компания Kaspi', default='', max_length=100, blank=True)
    kaspi_merchant_id = models.CharField('Merchant ID Kaspi', default='', max_length=100, blank=True)
    skip_merchant_id = models.CharField('Пропуск ID компаний', default = '', max_length=255, blank=True)

    def __str__(self):
        return self.email

    def get_skip_merchant_id(self):
        return self.skip_merchant_id.split(';')

    def save(self, *args, **kwargs):
        super(User, self).save(*args, **kwargs)

        if self.pricelist:
            file = self.pricelist.path
            data = pandas.read_excel(file, engine='openpyxl')

            path = shutil.copyfile(file, file.replace(self.pricelist.name, f'{self.pricelist.name}{timezone.now().strftime("%Y_%m_%d %H_%M_%S")}.xlsx'))
            with open(path, 'rb') as f:
                self.pricelist_xlsx = File(f, f'{self.pricelist.name}{timezone.now().strftime("%Y_%m_%d %H_%M_%S")}.xlsx')
                super(User, self).save(*args, **kwargs)

            wb = openpyxl.load_workbook(filename=self.pricelist_xlsx.path)
            ws = wb.active

            for row in range(2, ws.max_row + 1):
                if not ws.cell(row=row, column=1).value:
                    ws.cell(row=row, column=1, value='')
                    ws.cell(row=row, column=2, value='')
                    ws.cell(row=row, column=3, value='')
                    ws.cell(row=row, column=4, value='')
                    ws.cell(row=row, column=5, value='')

                ws.cell(row=row, column=6, value='')
                ws.cell(row=row, column=7, value='')
                ws.cell(row=row, column=8, value='')
                ws.cell(row=row, column=9, value='')
                ws.cell(row=row, column=10, value='')

            wb.save(self.pricelist_xlsx.path)

            # wb = openpyxl.load_workbook(file)
            # sheet = wb.active
            # rows = list(sheet.rows)
            # rows.remove(rows[0])

            if self.kaspi_company and self.kaspi_merchant_id:
                root = ET.Element("kaspi_catalog")

                root.set('date', f"{timezone.now}")
                root.set('xmlns', "kaspiShopping")
                root.set('xmlns:xsi', "http://www.w3.org/2001/XMLSchema-instance")
                root.set('xsi:schemaLocation', "kaspiShopping http://kaspi.kz/kaspishopping.xsd")

                company = ET.SubElement(root, 'company')
                company.text = self.kaspi_company
                merchantid = ET.SubElement(root, 'merchantid')
                merchantid.text = self.kaspi_merchant_id
                offers = ET.SubElement(root, 'offers')

                for index, row in data.iterrows():
                    try:
                        offer = ET.SubElement(offers, 'offer')
                        offer.set('sku', str(row['SKU']))

                        model = ET.SubElement(offer, 'model')
                        model.text = str(row['model']).replace('&', '&amp;').replace('"', '&quot;')
                        brand = ET.SubElement(offer, 'brand')
                        brand.text = str(row['brand'])

                        availabilities = ET.SubElement(offer, 'availabilities')
                        availability = ET.SubElement(availabilities, 'availability')
                        availability.set('storeId', 'PP1')
                        availability.set('available', row['PP1'])

                        price = ET.SubElement(offer, 'price')
                        price.text = str(row['price'])
                    except:
                        return None

                tree = ET.ElementTree(root)
                file_xml = os.path.join(settings.MEDIA_ROOT, f'pricelists_xml/{timezone.now().strftime("%Y_%m_%d %H_%M_%S")}.xml')
                tree.write(file_xml)
                with open(file_xml, 'rb') as f:
                    self.pricelist_xml = File(f, f'{timezone.now().strftime("%Y_%m_%d %H_%M_%S")}.xml')
                    super(User, self).save(*args, **kwargs)

    def check_active(self):
        if self.active:
            if not self.active_date > timezone.now():
                self.active = False
                self.save()
        return self.active

    def get_format_active_date(self):
        if self.active_date:
            return self.active_date.strftime("%Y-%m-%d")
        else:
            return None


class Subscription(models.Model):
    title = models.CharField('Название подписки', max_length=100)
    period = models.IntegerField('Период подписки')
    price = models.IntegerField('Цена подписки', default=0)
    free = models.BooleanField('Бесплатная', default=False)

    class Meta:
        verbose_name = 'Подписка'
        verbose_name_plural = 'Подписки'

    def __str__(self):
        return self.title

    def free(self):
        if self.price <= 0:
            return True
        return False
