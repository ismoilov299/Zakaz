import openpyxl
import os
from django.conf import settings
from bs4 import BeautifulSoup


def save_rows_in_file(file, rows):
    wb = openpyxl.load_workbook(filename=file)
    ws = wb.active

    for row in rows:
        ws.cell(row=int(row[0]), column=5, value=row[5])

    wb.save(file)


def parse_table(html):
    soup = BeautifulSoup(html, 'lxml')

    titles = soup.find('thead').find('tr').find_all('th')
    row = []
    for title in titles:
        row.append(title.text)

    rows = [row]
    rows_html = soup.find('tbody').find_all('tr')
    for row_html in rows_html:
        cols = row_html.find_all('td')
        row = []
        for col in cols:
            row.append(col.text)
        rows.append(row)
    return rows


def read_user_pricelist(user, offset, search):
    if not user.pricelist:
        return {}

    try:
        offset = int(offset) * 20
        file = os.path.join(settings.MEDIA_ROOT, user.pricelist.name)
        wookbook = openpyxl.load_workbook(file)
        worksheet = wookbook.active

        if search:
            return search_user_pricelist(search, worksheet, offset)

        rows = []
        offset_max = offset + 21
        if offset_max > worksheet.max_row:
            offset_max = worksheet.max_row
        for i in range(offset + 1, offset_max):
            cols = [f'{str(i + 1)}']
            for col in worksheet.iter_cols(1, worksheet.max_column):
                value = str(col[i].value)

                if value == 'None':
                    value = ''
                if len(value) > 50:
                    value = value[:50] + '...'

                cols.append(value)
            rows.append(cols)

        titles = ['1']
        for title in worksheet.iter_cols(1, worksheet.max_column):
            titles.append(str(title[0].value))

        data = {
            'titles': titles,
            'rows': rows,
            'amount_rows': worksheet.max_row,
        }
        return data
    except:
        return {}


def search_user_pricelist(search, worksheet, offset):
    search.replace('+', ' ')
    worksheet_rows = []
    for i in range(1, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if not col[i].value:
                continue
            if str(search) in str(col[i].value):
                cols = [f'{str(i + 1)}']
                for col in worksheet.iter_cols(1, worksheet.max_column):
                    value = str(col[i].value)
                    cols.append(value)
                worksheet_rows.append(cols)
                break
                
    offset_max = offset + 21
    if offset_max > len(worksheet_rows):
        offset_max = len(worksheet_rows)
    rows = []
    for i in range(offset, offset_max):
        cols = []
        for col in worksheet_rows[i]:
            value = col

            if value == 'None':
                value = ''
            if len(value) > 50:
                value = value[:50] + '...'

            cols.append(value)
        rows.append(cols)

    titles = ['1']
    for title in worksheet.iter_cols(1, worksheet.max_column):
        titles.append(str(title[0].value))

    data = {
        'titles': titles,
        'rows': rows,
        'amount_rows': len(worksheet_rows),
    }
    return data

